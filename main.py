# import webbrowser
# from openpyxl import load_workbook

# # Load the Excel file
# file_path = "C:/Users/ankit/Desktop/fiver.xlsx"
# wb = load_workbook(file_path)
# ws = wb.active

# # Specify Chrome executable path
# chrome_path = "C:/Program Files/Google/Chrome/Application/chrome.exe"  # Update this path with the actual path to chrome.exe

# # Iterate over rows in the Excel file
# for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3, max_col=3):  # Assuming site column is C
#     website_link = row[0].value
    
#     # Open website in Chrome browser
#     webbrowser.register('chrome', None, webbrowser.BackgroundBrowser(chrome_path))
#     webbrowser.get('chrome').open(website_link)

# # Close the Excel file (optional)
# wb.close()


import webbrowser
from openpyxl import load_workbook

# Load the Excel file
file_path = "C:/Users/ankit/Desktop/fiver.xlsx"
wb = load_workbook(file_path)
ws = wb.active

# Specify Brave executable path
brave_path = "C:/Program Files/BraveSoftware/Brave-Browser/Application/brave.exe"  # Update this path with the actual path to brave.exe

# Iterate over rows in the Excel file
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3, max_col=3):  # Assuming site column is C
    website_link = row[0].value
    
    # Open website in Brave browser
    webbrowser.register('brave', None, webbrowser.BackgroundBrowser(brave_path))
    webbrowser.get('brave').open(website_link)

# Close the Excel file (optional)
wb.close()
